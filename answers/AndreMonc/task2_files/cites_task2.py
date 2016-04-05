# !/usr/bin/env python
# encoding: utf-8

"""
My cites-list manipulation program
Created by Andre Moncrieff on 21 March 2016.
Copyright 2016 Andre E. Moncrieff. All rights reserved.
"""

import openpyxl
import re
import csv


def cites_list_cleaner(filename, column):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    cleaned_cites_list = []
    for rowNum in range(1, sheet.get_highest_row() + 1):
        cell_content = str(sheet.cell(row=rowNum, column=column).value)
        if re.search("[A-Z][a-z]+ [a-z]+", cell_content):
            capture = re.findall("([A-Z][a-z]+ [a-z]+)", cell_content)
            cleaned_cites_list.extend(capture)
        else:
            pass
    return cleaned_cites_list


def write_cleaned_cites_to_excel(cleaned_cites_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    rowNum = 1
    for species in cleaned_cites_list:
        ws.cell(row=rowNum, column=1).value = species
        rowNum += 1
    wb.save("clean_cites_list_of_Peruvian_birds.xlsx")


def make_trip_list(filename):
    wb2 = openpyxl.load_workbook(filename)
    trip_list_sheet = wb2.active
    hyp_trip_list = []
    for rowNum in range(1, trip_list_sheet.get_highest_row() + 1):
            cell_content = str(trip_list_sheet.cell(row=rowNum, column=1).value)
            if re.search("[A-Z][a-z]+ [a-z]+", cell_content):
                hyp_trip_list.append(cell_content)
    return hyp_trip_list


def trip_list_counted(some_trip_list):
    species_count = []
    for species in some_trip_list:
        total = 0
        for item in some_trip_list:
            if species == item:
                total += 1
        species_count.append([total])
    zip_dict = dict(zip(some_trip_list, species_count))
    return zip_dict


def add_cites_status_to_dict(trip_num_dict, cites_list):
    for key in trip_num_dict.keys():
        if key in cites_list:
            trip_num_dict[key].append("Cites")
        else:
            trip_num_dict[key].append("No Cites")
    return trip_num_dict


def write_summary_to_csv(final_dict):
    writer = csv.writer(open("cites_summary.csv", 'w'))
    writer.writerow(["Species", "Number Collected", "Cites Status"])
    for key, value in final_dict.items():
        writer.writerow([key, value[0], value[1]])


def main():
    raw_cites_filename = 'Raw_Peru_Avian_Cites_Dec_2014.xlsx'
    cleaned_cites_list = cites_list_cleaner(raw_cites_filename, 2)
    write_cleaned_cites_to_excel(cleaned_cites_list)
    trip_list_filename = 'Trip_List.xlsx'
    hypothetical_trip_list = make_trip_list(trip_list_filename)
    trip_num_dict = trip_list_counted(hypothetical_trip_list)
    final_dict = add_cites_status_to_dict(trip_num_dict, cleaned_cites_list)
    write_summary_to_csv(final_dict)


if __name__ == '__main__':
    main()
